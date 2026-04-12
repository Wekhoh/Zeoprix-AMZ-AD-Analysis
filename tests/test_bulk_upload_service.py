"""Tests for bulk_upload_service — Amazon Campaign Manager Excel generation."""

from io import BytesIO

from openpyxl import load_workbook

from backend.models import Campaign, KeywordAction, Marketplace
from backend.services.bulk_upload_service import generate_bulk_upload_excel


def _seed_actions(db_session):
    """Seed campaign + keyword actions for bulk upload testing."""
    mp = Marketplace(code="US", name="US", currency="USD")
    db_session.add(mp)
    db_session.flush()

    camp = Campaign(
        name="Test-SP-Harvest",
        ad_type="SP",
        targeting_type="auto",
        bidding_strategy="Fixed bids",
        status="Delivering",
        marketplace_id=mp.id,
    )
    db_session.add(camp)
    db_session.flush()

    actions = [
        KeywordAction(
            search_term="wireless charger",
            from_campaign_id=camp.id,
            from_campaign_name="Test-SP-Harvest",
            action_type="harvest_exact",
            target_bid=1.50,
        ),
        KeywordAction(
            search_term="phone stand",
            from_campaign_id=camp.id,
            from_campaign_name="Test-SP-Harvest",
            action_type="harvest_phrase",
            target_bid=0.80,
        ),
        KeywordAction(
            search_term="cheap knockoff",
            from_campaign_id=camp.id,
            from_campaign_name="Test-SP-Harvest",
            action_type="negate_exact",
        ),
        KeywordAction(
            search_term="free shipping only",
            from_campaign_id=camp.id,
            from_campaign_name="Test-SP-Harvest",
            action_type="negate_phrase",
        ),
    ]
    db_session.add_all(actions)
    db_session.commit()
    return camp


class TestBulkUploadGeneration:
    def test_generates_valid_xlsx(self, db_session):
        _seed_actions(db_session)
        excel_bytes = generate_bulk_upload_excel(db_session)

        assert isinstance(excel_bytes, bytes)
        assert len(excel_bytes) > 500

        wb = load_workbook(BytesIO(excel_bytes))
        assert "Harvest Keywords" in wb.sheetnames
        assert "Negative Keywords" in wb.sheetnames

    def test_harvest_sheet_has_correct_rows(self, db_session):
        _seed_actions(db_session)
        wb = load_workbook(BytesIO(generate_bulk_upload_excel(db_session)))
        ws = wb["Harvest Keywords"]

        # Header + 2 harvest actions
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        keywords = [r[4] for r in data_rows if r[0] == "Keyword"]
        assert "wireless charger" in keywords
        assert "phone stand" in keywords

    def test_harvest_match_types_correct(self, db_session):
        _seed_actions(db_session)
        wb = load_workbook(BytesIO(generate_bulk_upload_excel(db_session)))
        ws = wb["Harvest Keywords"]

        rows = {r[4]: r for r in ws.iter_rows(min_row=2, values_only=True) if r[0]}
        assert rows["wireless charger"][5] == "Exact"
        assert rows["phone stand"][5] == "Phrase"

    def test_harvest_bid_included(self, db_session):
        _seed_actions(db_session)
        wb = load_workbook(BytesIO(generate_bulk_upload_excel(db_session)))
        ws = wb["Harvest Keywords"]

        rows = {r[4]: r for r in ws.iter_rows(min_row=2, values_only=True) if r[0]}
        assert rows["wireless charger"][6] == 1.5

    def test_negate_sheet_has_correct_rows(self, db_session):
        _seed_actions(db_session)
        wb = load_workbook(BytesIO(generate_bulk_upload_excel(db_session)))
        ws = wb["Negative Keywords"]

        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        keywords = [r[3] for r in data_rows if r[0] and "Negative" in str(r[0])]
        assert "cheap knockoff" in keywords
        assert "free shipping only" in keywords

    def test_negate_match_types_correct(self, db_session):
        _seed_actions(db_session)
        wb = load_workbook(BytesIO(generate_bulk_upload_excel(db_session)))
        ws = wb["Negative Keywords"]

        rows = {r[3]: r for r in ws.iter_rows(min_row=2, values_only=True) if r[0]}
        assert rows["cheap knockoff"][4] == "Negative Exact"
        assert rows["free shipping only"][4] == "Negative Phrase"

    def test_filter_by_action_type(self, db_session):
        _seed_actions(db_session)
        excel_bytes = generate_bulk_upload_excel(db_session, action_types=["harvest_exact"])
        wb = load_workbook(BytesIO(excel_bytes))

        # Harvest sheet should only have harvest_exact
        ws1 = wb["Harvest Keywords"]
        data_rows = [r for r in ws1.iter_rows(min_row=2, values_only=True) if r[0] == "Keyword"]
        assert len(data_rows) == 1
        assert data_rows[0][4] == "wireless charger"

        # Negate sheet should be empty (no negate actions matched)
        ws2 = wb["Negative Keywords"]
        negate_rows = [
            r
            for r in ws2.iter_rows(min_row=2, values_only=True)
            if r[0] and "Negative" in str(r[0])
        ]
        assert len(negate_rows) == 0

    def test_empty_actions_produces_placeholder(self, db_session):
        """No KeywordAction records → sheets have placeholder text."""
        excel_bytes = generate_bulk_upload_excel(db_session)
        wb = load_workbook(BytesIO(excel_bytes))

        ws1 = wb["Harvest Keywords"]
        texts = [str(r[1]) for r in ws1.iter_rows(min_row=2, values_only=True)]
        assert any("无待收割" in t for t in texts)

    def test_safe_cell_applied_to_campaign_name(self, db_session):
        """Formula injection defense: campaign names starting with =
        must be escaped in the bulk upload Excel."""
        mp = Marketplace(code="US", name="US", currency="USD")
        db_session.add(mp)
        db_session.flush()
        camp = Campaign(
            name="=cmd|'/c calc'!A1",
            ad_type="SP",
            targeting_type="auto",
            bidding_strategy="Fixed bids",
            status="Delivering",
            marketplace_id=mp.id,
        )
        db_session.add(camp)
        db_session.flush()
        db_session.add(
            KeywordAction(
                search_term="test term",
                from_campaign_id=camp.id,
                from_campaign_name="=cmd|'/c calc'!A1",
                action_type="harvest_exact",
                target_bid=1.0,
            )
        )
        db_session.commit()

        wb = load_workbook(BytesIO(generate_bulk_upload_excel(db_session)))
        ws = wb["Harvest Keywords"]
        for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
            if row[1] and "cmd" in str(row[1]):
                assert str(row[1]).startswith("'"), f"Campaign name not escaped: {row[1]!r}"
                break
