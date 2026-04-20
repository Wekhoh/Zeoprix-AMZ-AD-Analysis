"""Tests for backend.services.report_service (H1).

Covers generate_excel_report():
- Return type + valid xlsx roundtrip via openpyxl.load_workbook.
- Expected 6-sheet structure with correct names + tab colors.
- KPI row values populated on 每日趋势 sheet.
- SUM/IF formulas present on the totals rows.
- Campaign rows sorted by spend descending on 广告活动对比 sheet.
- Empty state (no campaigns, no data) still produces a valid file.
"""

from io import BytesIO

from openpyxl import load_workbook

from backend.services.report_service import generate_excel_report

EXPECTED_SHEETS = [
    "报告摘要",
    "每日趋势",
    "广告活动对比",
    "展示位置分析",
    "交叉分析(花费)",
    "预警与建议",
]

EXPECTED_TAB_COLORS = {
    "报告摘要": "2563EB",
    "每日趋势": "10B981",
    "广告活动对比": "F59E0B",
    "展示位置分析": "8B5CF6",
    "交叉分析(花费)": "EF4444",
    "预警与建议": "DC2626",
}


class TestGenerateExcelReport:
    def test_returns_bytes(self, db_session, seed_campaign_data):
        result = generate_excel_report(db_session)
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_valid_xlsx_roundtrip(self, db_session, seed_campaign_data):
        data = generate_excel_report(db_session)
        wb = load_workbook(BytesIO(data))
        # Non-empty workbook loads successfully
        assert len(wb.sheetnames) > 0

    def test_has_six_expected_sheets(self, db_session, seed_campaign_data):
        data = generate_excel_report(db_session)
        wb = load_workbook(BytesIO(data))
        assert wb.sheetnames == EXPECTED_SHEETS

    def test_tab_colors_match_spec(self, db_session, seed_campaign_data):
        data = generate_excel_report(db_session)
        wb = load_workbook(BytesIO(data))
        for name, expected_color in EXPECTED_TAB_COLORS.items():
            actual = wb[name].sheet_properties.tabColor
            # tabColor is an openpyxl Color wrapper; .rgb gives "FFxxxxxx" or "xxxxxx"
            assert actual is not None, f"Sheet '{name}' missing tabColor"
            assert expected_color in str(actual.rgb), (
                f"Sheet '{name}' tabColor rgb={actual.rgb} missing {expected_color}"
            )

    def test_summary_sheet_has_title_and_date_range(self, db_session, seed_campaign_data):
        data = generate_excel_report(db_session, date_from="2025-11-10", date_to="2025-11-11")
        wb = load_workbook(BytesIO(data))
        ws = wb["报告摘要"]
        assert ws["A1"].value == "亚马逊广告智能追踪系统 — 分析报告"
        assert ws["A2"].value == "日期范围: 2025-11-10 ~ 2025-11-11"

    def test_summary_sheet_has_kpi_labels(self, db_session, seed_campaign_data):
        data = generate_excel_report(db_session)
        wb = load_workbook(BytesIO(data))
        ws = wb["报告摘要"]
        # KPI section subtitle at A4, table header row 5 (指标 / 数值)
        assert ws["A4"].value == "核心 KPI 概览"
        assert ws["A5"].value == "指标"
        assert ws["B5"].value == "数值"
        # Collect all label values in column A and verify key KPIs present
        labels = [ws.cell(row=r, column=1).value for r in range(6, 20)]
        assert "曝光量" in labels
        assert "花费" in labels
        assert "ROAS (广告回报)" in labels
        assert "ACOS (广告成本比)" in labels

    def test_daily_trend_sheet_has_header_and_data(self, db_session, seed_campaign_data):
        data = generate_excel_report(db_session)
        wb = load_workbook(BytesIO(data))
        ws = wb["每日趋势"]
        # Header row: 日期 + 10 KPI columns = 11 total
        assert ws.cell(row=1, column=1).value == "日期"
        assert ws.cell(row=1, column=2).value == "曝光量"
        assert ws.cell(row=1, column=4).value == "花费 ($)"
        # Two dates seeded (2025-11-10 and 2025-11-11) → rows 2 and 3
        date_values = {ws.cell(row=r, column=1).value for r in (2, 3)}
        assert "2025-11-10" in date_values
        assert "2025-11-11" in date_values

    def test_daily_trend_totals_row_uses_formulas(self, db_session, seed_campaign_data):
        data = generate_excel_report(db_session)
        wb = load_workbook(BytesIO(data))
        ws = wb["每日趋势"]
        # daily_data has 2 rows (2 dates), totals row at 4
        total_row = 4
        assert ws.cell(row=total_row, column=1).value == "合计"
        # SUM formula on spend column
        spend_formula = ws.cell(row=total_row, column=4).value
        assert isinstance(spend_formula, str)
        assert spend_formula.startswith("=SUM(")
        # CTR IF-guard formula on column 7
        ctr_formula = ws.cell(row=total_row, column=7).value
        assert isinstance(ctr_formula, str)
        assert "=IF(" in ctr_formula

    def test_campaign_sheet_sorted_by_spend_desc(self, db_session, seed_campaign_data):
        # seed_campaign_data creates: Test-SP-Auto (spend=60) and
        # Test-SP-Manual (spend=100). Higher-spend row must come first.
        data = generate_excel_report(db_session)
        wb = load_workbook(BytesIO(data))
        ws = wb["广告活动对比"]
        assert ws.cell(row=1, column=1).value == "广告活动"
        first_campaign = ws.cell(row=2, column=1).value
        second_campaign = ws.cell(row=3, column=1).value
        assert first_campaign == "Test-SP-Manual"  # spend 100
        assert second_campaign == "Test-SP-Auto"  # spend 60

    def test_placement_sheet_has_spend_percent_column(self, db_session, seed_campaign_data):
        data = generate_excel_report(db_session)
        wb = load_workbook(BytesIO(data))
        ws = wb["展示位置分析"]
        # Header includes "花费占比" as the last column (col 12)
        assert ws.cell(row=1, column=12).value == "花费占比"
        # Data row 2 — spend_pct formatted as percentage (0..1 float)
        pct_val = ws.cell(row=2, column=12).value
        assert pct_val is not None
        assert 0 <= pct_val <= 1
        assert ws.cell(row=2, column=12).number_format == "0.00%"

    def test_cross_sheet_pivot_has_campaign_and_placement_labels(
        self, db_session, seed_campaign_data
    ):
        data = generate_excel_report(db_session)
        wb = load_workbook(BytesIO(data))
        ws = wb["交叉分析(花费)"]
        # A1 is the corner header; row 1 cols 2+ are placements; col 1 rows 2+ are campaigns
        assert ws.cell(row=1, column=1).value == "广告活动 \\ 展示位置"
        # Collect placement headers from row 1 (exclude trailing 合计 if present)
        placement_headers = [
            ws.cell(row=1, column=c).value
            for c in range(2, ws.max_column)
            if ws.cell(row=1, column=c).value not in (None, "合计")
        ]
        assert "搜索顶部" in placement_headers

    def test_alerts_sheet_has_header_row(self, db_session, seed_campaign_data):
        data = generate_excel_report(db_session)
        wb = load_workbook(BytesIO(data))
        ws = wb["预警与建议"]
        assert ws.cell(row=1, column=1).value == "严重级别"
        assert ws.cell(row=1, column=2).value == "广告活动"
        assert ws.cell(row=1, column=3).value == "指标值"
        assert ws.cell(row=1, column=4).value == "建议"

    def test_empty_db_still_produces_valid_report(self, db_session):
        """With no campaigns seeded, the report should still build and load."""
        data = generate_excel_report(db_session)
        wb = load_workbook(BytesIO(data))
        assert wb.sheetnames == EXPECTED_SHEETS
        # Summary sheet still has title + KPI section header
        assert wb["报告摘要"]["A1"].value.startswith("亚马逊广告智能追踪系统")

    def test_date_filter_narrows_daily_trend(self, db_session, seed_campaign_data):
        # Filter to just 2025-11-10 → only one date row should appear
        data = generate_excel_report(db_session, date_from="2025-11-10", date_to="2025-11-10")
        wb = load_workbook(BytesIO(data))
        ws = wb["每日趋势"]
        dates = [
            ws.cell(row=r, column=1).value
            for r in range(2, ws.max_row + 1)
            if ws.cell(row=r, column=1).value and ws.cell(row=r, column=1).value != "合计"
        ]
        assert dates == ["2025-11-10"]
