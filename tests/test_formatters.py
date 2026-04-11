"""Tests for shared formatters — safe_cell injection defense + format helpers."""

from backend.services.formatters import (
    format_currency,
    format_float,
    format_int,
    format_percent,
    safe_cell,
)


class TestSafeCell:
    """Excel/CSV formula injection defense."""

    def test_none_passes_through(self):
        assert safe_cell(None) is None

    def test_int_passes_through(self):
        assert safe_cell(42) == 42

    def test_float_passes_through(self):
        assert safe_cell(3.14) == 3.14

    def test_bool_passes_through(self):
        assert safe_cell(True) is True

    def test_normal_string_passes_through(self):
        assert safe_cell("Campaign Name") == "Campaign Name"

    def test_empty_string_passes_through(self):
        assert safe_cell("") == ""

    def test_string_with_embedded_formula_chars_passes_through(self):
        # Only LEADING formula chars are dangerous
        assert safe_cell("Campaign=Alpha") == "Campaign=Alpha"
        assert safe_cell("Price: +10%") == "Price: +10%"

    def test_equals_sign_start_is_escaped(self):
        assert safe_cell("=SUM(A1:A10)") == "'=SUM(A1:A10)"

    def test_plus_start_is_escaped(self):
        assert safe_cell("+15%") == "'+15%"

    def test_minus_start_is_escaped(self):
        # Critical: negative number strings (e.g. from str(-42.5)) trigger formula
        assert safe_cell("-42.5") == "'-42.5"

    def test_at_start_is_escaped(self):
        # @SUM() is Lotus-1-2-3 formula syntax still honored by Excel
        assert safe_cell("@SUM(A1)") == "'@SUM(A1)"

    def test_tab_start_is_escaped(self):
        assert safe_cell("\tmalicious") == "'\tmalicious"

    def test_carriage_return_start_is_escaped(self):
        assert safe_cell("\revil") == "'\revil"

    def test_null_byte_start_is_escaped(self):
        assert safe_cell("\x00value") == "'\x00value"

    def test_realistic_amazon_injection_payload(self):
        # Example attack: malicious campaign name exfiltrates data
        payload = "=cmd|'/c calc'!A1"
        result = safe_cell(payload)
        assert result.startswith("'")
        assert "cmd" in result  # content preserved, just text-forced


class TestFormatCurrency:
    def test_none_returns_default(self):
        assert format_currency(None) == "-"

    def test_none_returns_custom_default(self):
        assert format_currency(None, default="N/A") == "N/A"

    def test_zero(self):
        assert format_currency(0) == "$0.00"

    def test_positive(self):
        assert format_currency(1234.56) == "$1,234.56"

    def test_large_number_comma_separator(self):
        assert format_currency(1234567.89) == "$1,234,567.89"

    def test_negative(self):
        assert format_currency(-100.5) == "$-100.50"

    def test_custom_symbol(self):
        assert format_currency(100, symbol="¥") == "¥100.00"

    def test_invalid_returns_default(self):
        assert format_currency("not a number") == "-"


class TestFormatPercent:
    def test_none_returns_default(self):
        assert format_percent(None) == "-"

    def test_zero(self):
        assert format_percent(0) == "0.00%"

    def test_ratio_to_percent(self):
        assert format_percent(0.1234) == "12.34%"

    def test_full_ratio(self):
        assert format_percent(1) == "100.00%"

    def test_custom_decimals(self):
        assert format_percent(0.12345, decimals=3) == "12.345%"

    def test_invalid_returns_default(self):
        assert format_percent("nope") == "-"


class TestFormatInt:
    def test_none_returns_default(self):
        assert format_int(None) == "-"

    def test_zero(self):
        assert format_int(0) == "0"

    def test_thousand_separator(self):
        assert format_int(12345) == "12,345"

    def test_large(self):
        assert format_int(1234567890) == "1,234,567,890"

    def test_float_truncates_to_int(self):
        assert format_int(42.9) == "42"

    def test_invalid_returns_default(self):
        assert format_int("x") == "-"


class TestFormatFloat:
    def test_none_returns_default(self):
        assert format_float(None) == "-"

    def test_default_2_decimals(self):
        assert format_float(3.14159) == "3.14"

    def test_custom_decimals(self):
        assert format_float(3.14159, decimals=4) == "3.1416"

    def test_zero_decimals(self):
        assert format_float(3.14, decimals=0) == "3"

    def test_invalid_returns_default(self):
        assert format_float("pi") == "-"


class TestReportServiceInjectionIntegration:
    """End-to-end: a malicious campaign name from DB must be text-escaped
    when passed through the Excel report generation pipeline.

    This is the 'wiring' test — unit tests prove safe_cell() works,
    this test proves report_service.py actually calls it.
    """

    def test_excel_report_escapes_malicious_campaign_name(self, db_session):
        from io import BytesIO

        from openpyxl import load_workbook

        from backend.models import Campaign, Marketplace, PlacementRecord
        from backend.services.report_service import generate_excel_report

        mp = Marketplace(code="US", name="US", currency="USD")
        db_session.add(mp)
        db_session.flush()

        # Realistic attack payload: opening the exported XLSX in Excel
        # without this defense would invoke a DDE formula
        malicious_name = "=cmd|'/c calc'!A1"
        evil = Campaign(
            name=malicious_name,
            ad_type="SP",
            targeting_type="auto",
            bidding_strategy="Fixed bids",
            status="Delivering",
            marketplace_id=mp.id,
        )
        db_session.add(evil)
        db_session.flush()
        db_session.add(
            PlacementRecord(
                date="2026-04-11",
                campaign_id=evil.id,
                placement_type="搜索顶部",
                impressions=100,
                clicks=10,
                spend=5.0,
                orders=1,
                sales=20.0,
            )
        )
        db_session.commit()

        # Generate Excel report and parse the bytes
        excel_bytes = generate_excel_report(db_session)
        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb["广告活动对比"]  # Sheet 3 from report_service

        # Find the row with our malicious campaign — its value should now
        # be prefixed with a leading single quote
        found = False
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            cell_value = row[0]
            if cell_value and "cmd" in str(cell_value):
                found = True
                assert str(cell_value).startswith("'"), (
                    f"Malicious campaign name was NOT escaped. Got: {cell_value!r}"
                )
                break
        assert found, "Could not find the malicious campaign in exported Excel"

    def test_excel_report_preserves_normal_campaign_names(self, db_session):
        """Regression: ordinary campaign names should NOT be mangled with ' prefix."""
        from io import BytesIO

        from openpyxl import load_workbook

        from backend.models import Campaign, Marketplace, PlacementRecord
        from backend.services.report_service import generate_excel_report

        mp = Marketplace(code="US", name="US", currency="USD")
        db_session.add(mp)
        db_session.flush()

        normal_name = "Normal Campaign 2026 Q2"
        c = Campaign(
            name=normal_name,
            ad_type="SP",
            targeting_type="auto",
            bidding_strategy="Fixed bids",
            status="Delivering",
            marketplace_id=mp.id,
        )
        db_session.add(c)
        db_session.flush()
        db_session.add(
            PlacementRecord(
                date="2026-04-11",
                campaign_id=c.id,
                placement_type="搜索顶部",
                impressions=100,
                clicks=10,
                spend=5.0,
                orders=1,
                sales=20.0,
            )
        )
        db_session.commit()

        excel_bytes = generate_excel_report(db_session)
        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb["广告活动对比"]

        found_name = None
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0] and "Normal" in str(row[0]):
                found_name = row[0]
                break
        assert found_name == normal_name, f"Normal name was unexpectedly modified: {found_name!r}"

    def test_pdf_report_generates_valid_bytes(self, db_session):
        """Smoke test: PDF report generation after formatters consolidation
        still produces a valid PDF byte stream with PDF magic header.

        Covers the B0-2b migration from local _fmt_* functions to
        backend.services.formatters — any format call signature mismatch
        or type error would crash generate_pdf_report.
        """
        from backend.models import Campaign, Marketplace, PlacementRecord
        from backend.services.pdf_report_service import generate_pdf_report

        mp = Marketplace(code="US", name="US", currency="USD")
        db_session.add(mp)
        db_session.flush()

        c = Campaign(
            name="Test PDF Campaign",
            ad_type="SP",
            targeting_type="auto",
            bidding_strategy="Fixed bids",
            status="Delivering",
            marketplace_id=mp.id,
        )
        db_session.add(c)
        db_session.flush()
        db_session.add(
            PlacementRecord(
                date="2026-04-11",
                campaign_id=c.id,
                placement_type="搜索顶部",
                impressions=1000,
                clicks=50,
                spend=25.0,
                orders=5,
                sales=150.0,
            )
        )
        db_session.commit()

        pdf_bytes = generate_pdf_report(db_session)
        assert isinstance(pdf_bytes, bytes)
        assert len(pdf_bytes) > 1000, "PDF should be non-trivial in size"
        assert pdf_bytes[:4] == b"%PDF", "Output must start with PDF magic header"

    def test_pdf_report_handles_none_kpi_values(self, db_session):
        """Edge case: all KPI values None (empty db) should render '-' not crash."""
        from backend.services.pdf_report_service import generate_pdf_report

        # Empty db — no campaigns, no placements
        pdf_bytes = generate_pdf_report(db_session)
        assert isinstance(pdf_bytes, bytes)
        assert pdf_bytes[:4] == b"%PDF"
