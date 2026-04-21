"""
Excel -> SQLite 迁移服务
从 亚马逊广告智能追踪系统_Ultimate.xlsx 迁移全部历史数据
"""

import tempfile
from pathlib import Path

from fastapi import UploadFile
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from backend.logging_config import get_logger
from backend.models import (
    AdGroup,
    AdGroupDailyRecord,
    Campaign,
    CampaignDailyRecord,
    ImportHistory,
    Marketplace,
    OperationLog,
    PlacementRecord,
    Product,
    ProductVariant,
)
from backend.schemas.import_result import ImportDetail, ImportResult
from backend.utils.campaign_parser import (
    extract_bidding_strategy_type,
    extract_default_bid,
    extract_variant_code,
    get_portfolio_name,
)

logger = get_logger("migration")

# Excel 工作表配置（与 data_importer.py 一致）
SHEET_CONFIG = {
    "placement": {"name": "展示位置", "start_row": 5},
    "campaign_data": {"name": "广告活动数据", "start_row": 6},
    "adgroup_data": {"name": "广告组数据", "start_row": 6},
    "campaign_log": {"name": "广告活动操作日志", "start_row": 6},
    "adgroup_log": {"name": "广告组操作日志", "start_row": 3},
}


def _seed_base_data(db: Session) -> tuple:
    """播种基础数据: 站点、产品、变体"""
    # 站点
    marketplace = db.query(Marketplace).filter_by(code="US").first()
    if not marketplace:
        marketplace = Marketplace(code="US", name="美国站", currency="USD")
        db.add(marketplace)
        db.flush()

    # 产品
    product = db.query(Product).filter_by(sku="ZP-TP01").first()
    if not product:
        product = Product(sku="ZP-TP01", name="ZEOPRIX Travel Pillow", category="旅行枕")
        db.add(product)
        db.flush()

    # 变体
    variants = {}
    for code, name in [("DBL", "双层枕"), ("BLK", "黑色款")]:
        v = db.query(ProductVariant).filter_by(product_id=product.id, variant_code=code).first()
        if not v:
            v = ProductVariant(
                product_id=product.id,
                variant_code=code,
                variant_name=name,
                marketplace_id=marketplace.id,
            )
            db.add(v)
            db.flush()
        variants[code] = v

    return marketplace, product, variants


def _get_or_create_campaign(
    db: Session,
    name: str,
    marketplace_id: int,
    variants: dict,
    bidding_strategy: str = "",
) -> Campaign:
    """获取或创建广告活动（迁移用）"""
    campaign = db.query(Campaign).filter_by(name=name, marketplace_id=marketplace_id).first()
    if campaign:
        return campaign

    variant_code = extract_variant_code(name)
    variant_id = variants[variant_code].id if variant_code and variant_code in variants else None
    strategy = bidding_strategy or extract_bidding_strategy_type(name)

    campaign = Campaign(
        name=name,
        marketplace_id=marketplace_id,
        variant_id=variant_id,
        ad_type="SP",
        targeting_type="auto",
        match_type="close",
        bidding_strategy=strategy,
        base_bid=extract_default_bid(name),
        portfolio=get_portfolio_name(name),
    )
    db.add(campaign)
    db.flush()

    # 1:1 广告组
    ad_group = AdGroup(
        campaign_id=campaign.id,
        name=name,
        default_bid=extract_default_bid(name),
    )
    db.add(ad_group)
    db.flush()

    return campaign


def _cell_str(val) -> str:
    """安全获取单元格字符串值"""
    if val is None:
        return ""
    return str(val).strip()[:10] if hasattr(val, "strftime") else str(val).strip()


def _cell_int(val) -> int:
    if val is None or val == "" or val == "—":
        return 0
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return 0


def _cell_float(val) -> float:
    if val is None or val == "" or val == "—":
        return 0.0
    try:
        return float(str(val).replace("$", "").replace(",", ""))
    except (ValueError, TypeError):
        return 0.0


async def migrate_excel_to_db(db: Session, file: UploadFile) -> ImportResult:
    """从 Excel 迁移全部历史数据到 SQLite"""
    details: list[ImportDetail] = []
    counts = {
        "placement": 0,
        "campaign_data": 0,
        "adgroup_data": 0,
        "campaign_log": 0,
        "adgroup_log": 0,
    }

    # 读取上传的 Excel 文件
    content = await file.read()
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.write(content)
    tmp.close()

    try:
        wb = load_workbook(tmp.name, data_only=True)

        # Step 1: 播种基础数据
        marketplace, product, variants = _seed_base_data(db)
        details.append(
            ImportDetail(message="基础数据播种完成 (US 站点, ZP-TP01 产品, DBL/BLK 变体)")
        )

        # Step 2: 迁移展示位置数据 (216 条)
        if SHEET_CONFIG["placement"]["name"] in wb.sheetnames:
            ws = wb[SHEET_CONFIG["placement"]["name"]]
            start = SHEET_CONFIG["placement"]["start_row"]

            for row in range(start, ws.max_row + 1):
                date_val = ws.cell(row=row, column=1).value
                campaign_name = ws.cell(row=row, column=2).value
                if not date_val or not campaign_name:
                    continue

                date_str = _cell_str(date_val)
                campaign = _get_or_create_campaign(db, str(campaign_name), marketplace.id, variants)

                placement_type = _cell_str(ws.cell(row=row, column=3).value)
                bidding_strategy = _cell_str(ws.cell(row=row, column=4).value)

                # 检查是否已存在
                exists = (
                    db.query(PlacementRecord)
                    .filter_by(
                        date=date_str,
                        campaign_id=campaign.id,
                        placement_type=placement_type,
                    )
                    .first()
                )
                if exists:
                    continue

                record = PlacementRecord(
                    date=date_str,
                    campaign_id=campaign.id,
                    placement_type=placement_type,
                    bidding_strategy=bidding_strategy,
                    impressions=_cell_int(ws.cell(row=row, column=5).value),
                    clicks=_cell_int(ws.cell(row=row, column=6).value),
                    spend=_cell_float(ws.cell(row=row, column=8).value),  # H 列
                    orders=_cell_int(ws.cell(row=row, column=10).value),  # J 列
                    sales=_cell_float(ws.cell(row=row, column=11).value),  # K 列
                )
                db.add(record)
                counts["placement"] += 1

            db.flush()
            details.append(ImportDetail(message=f"展示位置: 迁移 {counts['placement']} 条"))

        # Step 3: 迁移广告活动日数据
        if SHEET_CONFIG["campaign_data"]["name"] in wb.sheetnames:
            ws = wb[SHEET_CONFIG["campaign_data"]["name"]]
            start = SHEET_CONFIG["campaign_data"]["start_row"]

            for row in range(start, ws.max_row + 1):
                date_val = ws.cell(row=row, column=1).value
                campaign_name = ws.cell(row=row, column=2).value
                if not date_val or not campaign_name:
                    continue

                date_str = _cell_str(date_val)
                campaign = _get_or_create_campaign(
                    db,
                    str(campaign_name),
                    marketplace.id,
                    variants,
                    _cell_str(ws.cell(row=row, column=5).value),
                )

                exists = (
                    db.query(CampaignDailyRecord)
                    .filter_by(
                        date=date_str,
                        campaign_id=campaign.id,
                    )
                    .first()
                )
                if exists:
                    continue

                record = CampaignDailyRecord(
                    date=date_str,
                    campaign_id=campaign.id,
                    status=_cell_str(ws.cell(row=row, column=3).value),
                    budget=_cell_float(ws.cell(row=row, column=6).value) or None,
                    impressions=_cell_int(ws.cell(row=row, column=7).value),
                    clicks=_cell_int(ws.cell(row=row, column=8).value),
                    spend=_cell_float(ws.cell(row=row, column=10).value),  # J 列
                    orders=_cell_int(ws.cell(row=row, column=12).value),  # L 列
                    sales=_cell_float(ws.cell(row=row, column=13).value),  # M 列
                )
                db.add(record)
                counts["campaign_data"] += 1

            db.flush()
            details.append(ImportDetail(message=f"广告活动数据: 迁移 {counts['campaign_data']} 条"))

        # Step 4: 迁移广告组日数据
        if SHEET_CONFIG["adgroup_data"]["name"] in wb.sheetnames:
            ws = wb[SHEET_CONFIG["adgroup_data"]["name"]]
            start = SHEET_CONFIG["adgroup_data"]["start_row"]

            for row in range(start, ws.max_row + 1):
                date_val = ws.cell(row=row, column=1).value
                campaign_name = ws.cell(row=row, column=2).value
                if not date_val or not campaign_name:
                    continue

                date_str = _cell_str(date_val)
                campaign = _get_or_create_campaign(db, str(campaign_name), marketplace.id, variants)
                ad_group = db.query(AdGroup).filter_by(campaign_id=campaign.id).first()

                if not ad_group:
                    continue

                exists = (
                    db.query(AdGroupDailyRecord)
                    .filter_by(
                        date=date_str,
                        ad_group_id=ad_group.id,
                    )
                    .first()
                )
                if exists:
                    continue

                record = AdGroupDailyRecord(
                    date=date_str,
                    ad_group_id=ad_group.id,
                    campaign_id=campaign.id,
                    default_bid=_cell_float(ws.cell(row=row, column=5).value) or None,
                    impressions=_cell_int(ws.cell(row=row, column=7).value),
                    clicks=_cell_int(ws.cell(row=row, column=8).value),
                    spend=_cell_float(ws.cell(row=row, column=10).value),
                    orders=_cell_int(ws.cell(row=row, column=12).value),
                    sales=_cell_float(ws.cell(row=row, column=13).value),
                )
                db.add(record)
                counts["adgroup_data"] += 1

            db.flush()
            details.append(ImportDetail(message=f"广告组数据: 迁移 {counts['adgroup_data']} 条"))

        # Step 5: 迁移广告活动操作日志
        if SHEET_CONFIG["campaign_log"]["name"] in wb.sheetnames:
            ws = wb[SHEET_CONFIG["campaign_log"]["name"]]
            start = SHEET_CONFIG["campaign_log"]["start_row"]

            for row in range(start, ws.max_row + 1):
                date_val = ws.cell(row=row, column=1).value
                campaign_name = ws.cell(row=row, column=5).value
                if not date_val or not campaign_name:
                    continue

                date_str = _cell_str(date_val)
                time_str = _cell_str(ws.cell(row=row, column=2).value)
                campaign = _get_or_create_campaign(db, str(campaign_name), marketplace.id, variants)

                change_type = _cell_str(ws.cell(row=row, column=8).value)
                from_val = _cell_str(ws.cell(row=row, column=9).value)
                to_val = _cell_str(ws.cell(row=row, column=10).value)

                # 检查重复（使用复合键含 level_type）
                exists = (
                    db.query(OperationLog)
                    .filter_by(
                        date=date_str,
                        time=time_str,
                        level_type="campaign",
                        campaign_id=campaign.id,
                        change_type=change_type,
                        from_value=from_val,
                        to_value=to_val,
                    )
                    .first()
                )
                if exists:
                    continue

                log_record = OperationLog(
                    date=date_str,
                    time=time_str,
                    operator=_cell_str(ws.cell(row=row, column=3).value),
                    level_type="campaign",
                    campaign_id=campaign.id,
                    operation_type=_cell_str(ws.cell(row=row, column=7).value),
                    change_type=change_type,
                    from_value=from_val,
                    to_value=to_val,
                )
                try:
                    db.add(log_record)
                    db.flush()
                    counts["campaign_log"] += 1
                except Exception as e:
                    db.rollback()
                    logger.warning(f"Campaign log migration skipped: {e}")

            details.append(
                ImportDetail(message=f"广告活动操作日志: 迁移 {counts['campaign_log']} 条")
            )

        # Step 6: 迁移广告组操作日志
        if SHEET_CONFIG["adgroup_log"]["name"] in wb.sheetnames:
            ws = wb[SHEET_CONFIG["adgroup_log"]["name"]]
            start = SHEET_CONFIG["adgroup_log"]["start_row"]

            for row in range(start, ws.max_row + 1):
                date_val = ws.cell(row=row, column=1).value
                campaign_name = ws.cell(row=row, column=5).value
                if not date_val or not campaign_name:
                    continue

                date_str = _cell_str(date_val)
                time_str = _cell_str(ws.cell(row=row, column=2).value)
                campaign = _get_or_create_campaign(db, str(campaign_name), marketplace.id, variants)
                ad_group = db.query(AdGroup).filter_by(campaign_id=campaign.id).first()

                change_type = _cell_str(ws.cell(row=row, column=8).value)
                from_val = _cell_str(ws.cell(row=row, column=9).value)
                to_val = _cell_str(ws.cell(row=row, column=10).value)

                exists = (
                    db.query(OperationLog)
                    .filter_by(
                        date=date_str,
                        time=time_str,
                        level_type="ad_group",
                        campaign_id=campaign.id,
                        change_type=change_type,
                        from_value=from_val,
                        to_value=to_val,
                    )
                    .first()
                )
                if exists:
                    continue

                log_record = OperationLog(
                    date=date_str,
                    time=time_str,
                    operator=_cell_str(ws.cell(row=row, column=3).value),
                    level_type="ad_group",
                    campaign_id=campaign.id,
                    ad_group_id=ad_group.id if ad_group else None,
                    operation_type=_cell_str(ws.cell(row=row, column=7).value),
                    change_type=change_type,
                    from_value=from_val,
                    to_value=to_val,
                )
                try:
                    db.add(log_record)
                    db.flush()
                    counts["adgroup_log"] += 1
                except Exception as e:
                    db.rollback()
                    logger.warning(f"Ad group log migration skipped: {e}")

            details.append(ImportDetail(message=f"广告组操作日志: 迁移 {counts['adgroup_log']} 条"))

        # 更新广告活动状态
        _update_campaign_statuses(db)

        db.commit()
        wb.close()

        total = sum(counts.values())
        details.append(ImportDetail(message=f"迁移完成! 总计 {total} 条记录"))

        # 记录导入历史
        db.add(
            ImportHistory(
                import_type="migration",
                file_name=file.filename,
                records_imported=total,
                status="success",
            )
        )
        db.commit()

        return ImportResult(imported=total, details=details)

    except Exception as e:
        db.rollback()
        return ImportResult(
            errors=1,
            details=[ImportDetail(message=f"迁移失败: {e}", level="error")],
        )
    finally:
        Path(tmp.name).unlink(missing_ok=True)


def _update_campaign_statuses(db: Session):
    """从操作日志推断广告活动当前状态"""
    valid_statuses = ["Paused", "Delivering", "Enabled", "Archived"]

    campaigns = db.query(Campaign).all()
    for campaign in campaigns:
        # 找最后一条状态变更记录
        last_status_log = (
            db.query(OperationLog)
            .filter(
                OperationLog.campaign_id == campaign.id,
                OperationLog.change_type.contains("Campaign status"),
            )
            .order_by(OperationLog.date.desc(), OperationLog.time.desc())
            .first()
        )

        if last_status_log and any(s in str(last_status_log.to_value) for s in valid_statuses):
            campaign.status = last_status_log.to_value
            campaign.status_updated_at = last_status_log.date
