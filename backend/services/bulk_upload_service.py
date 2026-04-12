"""Amazon Campaign Manager Bulk Upload Excel generation.

Converts KeywordAction records into an Excel file matching Amazon's
Sponsored Products bulk upload format. The operator downloads this file
and uploads it directly to Seller Central → Campaign Manager → Bulk
Operations, avoiding manual keyword-by-keyword entry.

Format reference: Amazon Seller Central → Campaign Manager → Bulk
Operations → Download template. Columns are the minimum required set
for keyword operations (add exact keywords + add negative keywords).
"""

from io import BytesIO
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy.orm import Session

from backend.models import KeywordAction
from backend.services.formatters import safe_cell

# Amazon Bulk Upload column headers (SP Keyword operations)
HARVEST_HEADERS = [
    "Record Type",
    "Campaign Name",
    "Ad Group Name",
    "State",
    "Keyword or Product Targeting",
    "Match Type",
    "Max Bid",
]

NEGATE_HEADERS = [
    "Record Type",
    "Campaign Name",
    "State",
    "Keyword or Product Targeting",
    "Match Type",
]

# Styles (consistent with report_service.py)
_HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
_HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
_BODY_FONT = Font(name="Arial", size=10)
_THIN_BORDER = Border(
    left=Side(style="thin", color="E5E7EB"),
    right=Side(style="thin", color="E5E7EB"),
    top=Side(style="thin", color="E5E7EB"),
    bottom=Side(style="thin", color="E5E7EB"),
)

# Map our action_type values to Amazon's Match Type column
_MATCH_TYPE_MAP = {
    "harvest_exact": "Exact",
    "harvest_phrase": "Phrase",
    "negate_exact": "Negative Exact",
    "negate_phrase": "Negative Phrase",
}

# Map to Amazon's Record Type
_RECORD_TYPE_MAP = {
    "harvest_exact": "Keyword",
    "harvest_phrase": "Keyword",
    "negate_exact": "Campaign Negative Keyword",
    "negate_phrase": "Campaign Negative Keyword",
}


def _style_header(ws, row, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _THIN_BORDER


def _auto_width(ws, min_w=10, max_w=40):
    for col in ws.columns:
        letter = None
        max_len = 0
        for cell in col:
            if hasattr(cell, "column_letter"):
                letter = cell.column_letter
            val = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, len(val.encode("utf-8")))
        if letter:
            ws.column_dimensions[letter].width = max(min(max_len * 0.8 + 2, max_w), min_w)


def generate_bulk_upload_excel(
    db: Session,
    action_types: Optional[list[str]] = None,
) -> bytes:
    """Query KeywordAction records and generate an Amazon Bulk Upload Excel.

    Args:
        db: SQLAlchemy session
        action_types: filter by action_type (e.g., ["harvest_exact", "negate_exact"]).
            If None, exports all recorded actions.

    Returns:
        Excel file as bytes, ready for StreamingResponse or file write.
    """
    q = db.query(KeywordAction).order_by(KeywordAction.created_at.desc())
    if action_types:
        q = q.filter(KeywordAction.action_type.in_(action_types))
    actions = q.all()

    # Split into harvest (positive) and negate (negative) groups
    harvest_actions = [a for a in actions if a.action_type.startswith("harvest")]
    negate_actions = [a for a in actions if a.action_type.startswith("negate")]

    wb = Workbook()

    # === Sheet 1: Harvest Keywords (add new exact/phrase keywords)
    ws1 = wb.active
    ws1.title = "Harvest Keywords"
    ws1.sheet_properties.tabColor = "10B981"

    ws1.append(HARVEST_HEADERS)
    _style_header(ws1, 1, len(HARVEST_HEADERS))

    for a in harvest_actions:
        record_type = _RECORD_TYPE_MAP.get(a.action_type, "Keyword")
        match_type = _MATCH_TYPE_MAP.get(a.action_type, "Exact")
        row = [
            record_type,
            safe_cell(a.from_campaign_name or ""),
            "",  # Ad Group Name — operator fills manually
            "enabled",
            safe_cell(a.search_term),
            match_type,
            a.target_bid if a.target_bid else "",
        ]
        ws1.append(row)
        for col in range(1, len(HARVEST_HEADERS) + 1):
            ws1.cell(row=ws1.max_row, column=col).font = _BODY_FONT
            ws1.cell(row=ws1.max_row, column=col).border = _THIN_BORDER

    if not harvest_actions:
        ws1.append(["", "（无待收割关键词）", "", "", "", "", ""])

    _auto_width(ws1)

    # === Sheet 2: Negative Keywords (campaign-level negative)
    ws2 = wb.create_sheet("Negative Keywords")
    ws2.sheet_properties.tabColor = "EF4444"

    ws2.append(NEGATE_HEADERS)
    _style_header(ws2, 1, len(NEGATE_HEADERS))

    for a in negate_actions:
        record_type = _RECORD_TYPE_MAP.get(a.action_type, "Campaign Negative Keyword")
        match_type = _MATCH_TYPE_MAP.get(a.action_type, "Negative Exact")
        row = [
            record_type,
            safe_cell(a.from_campaign_name or ""),
            "enabled",
            safe_cell(a.search_term),
            match_type,
        ]
        ws2.append(row)
        for col in range(1, len(NEGATE_HEADERS) + 1):
            ws2.cell(row=ws2.max_row, column=col).font = _BODY_FONT
            ws2.cell(row=ws2.max_row, column=col).border = _THIN_BORDER

    if not negate_actions:
        ws2.append(["", "（无待否定关键词）", "", "", ""])

    _auto_width(ws2)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# Amazon action_type → Bulk Upload Record Type + action column mapping
_SUGGESTION_ACTION_MAP = {
    "flag_pause": {"record_type": "Campaign", "state": "paused", "note": "暂停广告活动"},
    "suggest_bid_decrease": {
        "record_type": "Campaign",
        "state": "",
        "note": "降低竞价（需手动填写新值）",
    },
    "suggest_bid_increase": {
        "record_type": "Campaign",
        "state": "",
        "note": "提高竞价（需手动填写新值）",
    },
    "suggest_budget_increase": {
        "record_type": "Campaign",
        "state": "",
        "note": "增加日预算（需手动填写新值）",
    },
    "suggest_negative": {
        "record_type": "Campaign Negative Keyword",
        "state": "enabled",
        "note": "添加否定关键词",
    },
}

SUGGESTION_HEADERS = [
    "Record Type",
    "Campaign Name",
    "State",
    "Daily Budget",
    "Bid",
    "参考: 建议操作",
    "参考: 触发指标值",
    "参考: 规则名称",
]


def generate_suggestion_bulk_upload(suggestions: list[dict]) -> bytes:
    """Generate Amazon Bulk Upload Excel from rule evaluation suggestions.

    Columns marked "参考:" are helper columns the operator should DELETE
    before uploading to Amazon. They provide context for filling in the
    actual bid/budget values.

    Args:
        suggestions: list of dicts from evaluate_rules / get_rule_results.
            Each dict must have: campaign_name, action_type, recommended_action,
            triggered_value, rule_name.
    """
    wb = Workbook()

    # Separate actionable vs informational suggestions
    actionable = [s for s in suggestions if s.get("action_type") in _SUGGESTION_ACTION_MAP]
    informational = [s for s in suggestions if s.get("action_type") not in _SUGGESTION_ACTION_MAP]

    # === Sheet 1: Actionable (can map to bulk upload)
    ws1 = wb.active
    ws1.title = "Bulk Upload 操作"
    ws1.sheet_properties.tabColor = "2563EB"

    ws1.append(SUGGESTION_HEADERS)
    _style_header(ws1, 1, len(SUGGESTION_HEADERS))

    for s in actionable:
        mapping = _SUGGESTION_ACTION_MAP[s["action_type"]]
        row = [
            mapping["record_type"],
            safe_cell(s.get("campaign_name", "")),
            mapping["state"] or "（需填写）",
            "",  # Daily Budget — operator fills
            "",  # Bid — operator fills
            s.get("recommended_action", ""),
            str(s.get("triggered_value", "")),
            s.get("rule_name", ""),
        ]
        ws1.append(row)
        for col in range(1, len(SUGGESTION_HEADERS) + 1):
            ws1.cell(row=ws1.max_row, column=col).font = _BODY_FONT
            ws1.cell(row=ws1.max_row, column=col).border = _THIN_BORDER

    if not actionable:
        ws1.append(["", "（无可操作的建议）", "", "", "", "", "", ""])

    _auto_width(ws1)

    # === Sheet 2: Informational (reminders, diagnostics — no bulk upload action)
    ws2 = wb.create_sheet("参考信息")
    ws2.sheet_properties.tabColor = "6B7280"

    info_headers = ["广告活动", "建议操作", "触发指标值", "规则名称"]
    ws2.append(info_headers)
    _style_header(ws2, 1, len(info_headers))

    for s in informational:
        row = [
            safe_cell(s.get("campaign_name", "")),
            s.get("recommended_action", ""),
            str(s.get("triggered_value", "")),
            s.get("rule_name", ""),
        ]
        ws2.append(row)
        for col in range(1, len(info_headers) + 1):
            ws2.cell(row=ws2.max_row, column=col).font = _BODY_FONT
            ws2.cell(row=ws2.max_row, column=col).border = _THIN_BORDER

    if not informational:
        ws2.append(["", "（无参考信息）", "", ""])

    _auto_width(ws2)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
