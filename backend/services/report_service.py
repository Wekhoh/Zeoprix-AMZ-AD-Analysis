"""专业级 Excel 报告生成服务"""

from io import BytesIO
from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from backend.services.summary_service import (
    summary_by_date,
    summary_by_campaign,
    summary_by_placement,
    dashboard_overview,
)

# ============================================================
# Styles
# ============================================================
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
SUBHEADER_FONT = Font(name="Arial", bold=True, size=11)
BODY_FONT = Font(name="Arial", size=10)
TITLE_FONT = Font(name="Arial", bold=True, size=14, color="1F2937")
SUBTITLE_FONT = Font(name="Arial", bold=True, size=12, color="374151")
GOOD_FILL = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
BAD_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="E5E7EB"),
    right=Side(style="thin", color="E5E7EB"),
    top=Side(style="thin", color="E5E7EB"),
    bottom=Side(style="thin", color="E5E7EB"),
)

# Number formats
FMT_NUM = "#,##0"
FMT_CURRENCY = "$#,##0.00"
FMT_PCT = "0.00%"
FMT_DECIMAL = "0.00"


def _style_header(ws, row, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def _style_row(ws, row, col_count, font=None):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        if font:
            cell.font = font
        cell.border = THIN_BORDER


def _auto_width(ws, min_width=8, max_width=35):
    for col in ws.columns:
        max_len = 0
        col_letter = None
        for cell in col:
            if hasattr(cell, "column_letter"):
                col_letter = cell.column_letter
            val = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, len(val.encode("utf-8")))
        if col_letter:
            ws.column_dimensions[col_letter].width = max(
                min(max_len * 0.8 + 2, max_width), min_width
            )


def _apply_kpi_formats(ws, row, col_start, fmt_map):
    """Apply number formats to KPI columns. fmt_map: {col_offset: format}"""
    for offset, fmt in fmt_map.items():
        cell = ws.cell(row=row, column=col_start + offset)
        if cell.value is not None:
            cell.number_format = fmt


# KPI column order: impressions, clicks, spend, orders, sales, ctr, cpc, roas, acos, cvr
KPI_HEADERS = [
    "曝光量",
    "点击量",
    "花费 ($)",
    "订单",
    "销售额 ($)",
    "CTR",
    "CPC ($)",
    "ROAS",
    "ACOS",
    "CVR",
]
KPI_KEYS = [
    "impressions",
    "clicks",
    "spend",
    "orders",
    "sales",
    "ctr",
    "cpc",
    "roas",
    "acos",
    "cvr",
]
# Format map: offset from first KPI column → format
KPI_FMT = {
    0: FMT_NUM,
    1: FMT_NUM,
    2: FMT_CURRENCY,
    3: FMT_NUM,
    4: FMT_CURRENCY,
    5: FMT_PCT,
    6: FMT_CURRENCY,
    7: FMT_DECIMAL,
    8: FMT_PCT,
    9: FMT_PCT,
}


def _write_kpi_row(ws, row, data, kpi_col_start):
    for i, key in enumerate(KPI_KEYS):
        cell = ws.cell(row=row, column=kpi_col_start + i, value=data.get(key))
        cell.number_format = KPI_FMT[i]
        cell.font = BODY_FONT
        cell.border = THIN_BORDER


def generate_excel_report(
    db: Session,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
) -> bytes:
    wb = Workbook()

    overview = dashboard_overview(db, date_from, date_to)
    daily_data = summary_by_date(db, date_from, date_to)
    campaign_data = summary_by_campaign(db, date_from, date_to)
    placement_data = summary_by_placement(db, date_from, date_to)

    # ============================================================
    # Sheet 1: 报告摘要（Dashboard Overview）
    # ============================================================
    ws = wb.active
    ws.title = "报告摘要"
    ws.sheet_properties.tabColor = "2563EB"

    # Title
    ws.merge_cells("A1:F1")
    ws["A1"] = "亚马逊广告智能追踪系统 — 分析报告"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    date_range = f"日期范围: {date_from or '全部'} ~ {date_to or '全部'}"
    ws.merge_cells("A2:F2")
    ws["A2"] = date_range
    ws["A2"].font = Font(name="Arial", size=10, color="6B7280")

    # KPI Summary Table
    ws["A4"] = "核心 KPI 概览"
    ws["A4"].font = SUBTITLE_FONT

    kpi = overview["kpi"]
    kpi_rows = [
        ("曝光量", kpi.get("impressions", 0), FMT_NUM),
        ("点击量", kpi.get("clicks", 0), FMT_NUM),
        ("花费", kpi.get("spend", 0), FMT_CURRENCY),
        ("订单数", kpi.get("orders", 0), FMT_NUM),
        ("销售额", kpi.get("sales", 0), FMT_CURRENCY),
        ("CTR (点击率)", kpi.get("ctr"), FMT_PCT),
        ("CPC (点击成本)", kpi.get("cpc"), FMT_CURRENCY),
        ("ROAS (广告回报)", kpi.get("roas"), FMT_DECIMAL),
        ("ACOS (广告成本比)", kpi.get("acos"), FMT_PCT),
        ("CVR (转化率)", kpi.get("cvr"), FMT_PCT),
    ]

    ws.append(["指标", "数值"])
    _style_header(ws, 5, 2)
    for label, val, fmt in kpi_rows:
        row_num = ws.max_row + 1
        ws.cell(row=row_num, column=1, value=label).font = BODY_FONT
        ws.cell(row=row_num, column=1).border = THIN_BORDER
        cell = ws.cell(row=row_num, column=2, value=val)
        cell.number_format = fmt
        cell.font = BODY_FONT
        cell.border = THIN_BORDER

    # Campaign status summary
    row_start = ws.max_row + 2
    ws.cell(row=row_start, column=1, value="广告活动状态分布").font = SUBTITLE_FONT
    row_start += 1
    ws.append(["状态", "数量"])
    _style_header(ws, row_start, 2)
    for status, count in overview.get("status_counts", {}).items():
        r = ws.max_row + 1
        ws.cell(row=r, column=1, value=status).font = BODY_FONT
        ws.cell(row=r, column=1).border = THIN_BORDER
        ws.cell(row=r, column=2, value=count).font = BODY_FONT
        ws.cell(row=r, column=2).border = THIN_BORDER

    _auto_width(ws)

    # ============================================================
    # Sheet 2: 每日趋势
    # ============================================================
    ws2 = wb.create_sheet("每日趋势")
    ws2.sheet_properties.tabColor = "10B981"

    headers = ["日期"] + KPI_HEADERS
    ws2.append(headers)
    _style_header(ws2, 1, len(headers))

    for i, row in enumerate(daily_data):
        r = i + 2
        ws2.cell(row=r, column=1, value=row.get("date")).font = BODY_FONT
        ws2.cell(row=r, column=1).border = THIN_BORDER
        _write_kpi_row(ws2, r, row, 2)

    # Add daily totals row
    if daily_data:
        total_row = len(daily_data) + 2
        ws2.cell(row=total_row, column=1, value="合计").font = Font(
            name="Arial", bold=True, size=10
        )
        ws2.cell(row=total_row, column=1).border = THIN_BORDER
        for col_idx in [2, 3, 4, 5, 6]:  # impressions, clicks, spend, orders, sales
            col_letter = get_column_letter(col_idx)
            cell = ws2.cell(row=total_row, column=col_idx)
            cell.value = f"=SUM({col_letter}2:{col_letter}{total_row - 1})"
            cell.font = Font(name="Arial", bold=True, size=10)
            cell.border = THIN_BORDER
            cell.number_format = KPI_FMT[col_idx - 2]
        # CTR = clicks/impressions
        ws2.cell(
            row=total_row,
            column=7,
            value=f"=IF(B{total_row}=0,0,C{total_row}/B{total_row})",
        ).number_format = FMT_PCT
        # CPC = spend/clicks
        ws2.cell(
            row=total_row,
            column=8,
            value=f"=IF(C{total_row}=0,0,D{total_row}/C{total_row})",
        ).number_format = FMT_CURRENCY
        # ROAS = sales/spend
        ws2.cell(
            row=total_row,
            column=9,
            value=f"=IF(D{total_row}=0,0,F{total_row}/D{total_row})",
        ).number_format = FMT_DECIMAL
        # ACOS = spend/sales
        ws2.cell(
            row=total_row,
            column=10,
            value=f"=IF(F{total_row}=0,0,D{total_row}/F{total_row})",
        ).number_format = FMT_PCT
        # CVR = orders/clicks
        ws2.cell(
            row=total_row,
            column=11,
            value=f"=IF(C{total_row}=0,0,E{total_row}/C{total_row})",
        ).number_format = FMT_PCT

        for col_idx in range(7, 12):
            ws2.cell(row=total_row, column=col_idx).font = Font(name="Arial", bold=True, size=10)
            ws2.cell(row=total_row, column=col_idx).border = THIN_BORDER

    _auto_width(ws2)

    # ============================================================
    # Sheet 3: 广告活动对比分析（核心新增！）
    # ============================================================
    ws3 = wb.create_sheet("广告活动对比")
    ws3.sheet_properties.tabColor = "F59E0B"

    headers3 = ["广告活动", "状态", "数据范围"] + KPI_HEADERS
    ws3.append(headers3)
    _style_header(ws3, 1, len(headers3))

    # Sort by spend descending
    sorted_campaigns = sorted(campaign_data, key=lambda x: x.get("spend", 0), reverse=True)

    for i, row in enumerate(sorted_campaigns):
        r = i + 2
        ws3.cell(row=r, column=1, value=row.get("campaign_name")).font = BODY_FONT
        ws3.cell(row=r, column=1).border = THIN_BORDER

        status_cell = ws3.cell(row=r, column=2, value=row.get("status"))
        status_cell.font = BODY_FONT
        status_cell.border = THIN_BORDER
        if row.get("status") == "Paused":
            status_cell.fill = PatternFill(
                start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"
            )

        date_range_str = f"{row.get('first_date', '')} ~ {row.get('last_date', '')}"
        ws3.cell(row=r, column=3, value=date_range_str).font = BODY_FONT
        ws3.cell(row=r, column=3).border = THIN_BORDER

        _write_kpi_row(ws3, r, row, 4)

        # Highlight high ACOS (>50%) in red
        acos_cell = ws3.cell(row=r, column=12)  # ACOS column (4+8)
        if row.get("acos") and row["acos"] > 0.5:
            acos_cell.fill = BAD_FILL
        # Highlight good ROAS (>1) in green
        roas_cell = ws3.cell(row=r, column=11)  # ROAS column (4+7)
        if row.get("roas") and row["roas"] > 1:
            roas_cell.fill = GOOD_FILL

    # Totals
    if sorted_campaigns:
        total_row = len(sorted_campaigns) + 2
        ws3.cell(row=total_row, column=1, value="合计").font = Font(
            name="Arial", bold=True, size=10
        )
        for col_offset in range(5):
            col_idx = 4 + col_offset
            col_letter = get_column_letter(col_idx)
            cell = ws3.cell(row=total_row, column=col_idx)
            cell.value = f"=SUM({col_letter}2:{col_letter}{total_row - 1})"
            cell.font = Font(name="Arial", bold=True, size=10)
            cell.border = THIN_BORDER
            cell.number_format = KPI_FMT[col_offset]

    _auto_width(ws3)

    # ============================================================
    # Sheet 4: 展示位置分析
    # ============================================================
    ws4 = wb.create_sheet("展示位置分析")
    ws4.sheet_properties.tabColor = "8B5CF6"

    headers4 = ["展示位置"] + KPI_HEADERS + ["花费占比"]
    ws4.append(headers4)
    _style_header(ws4, 1, len(headers4))

    total_spend = sum(p.get("spend", 0) for p in placement_data) or 1

    for i, row in enumerate(placement_data):
        r = i + 2
        ws4.cell(row=r, column=1, value=row.get("placement_type")).font = BODY_FONT
        ws4.cell(row=r, column=1).border = THIN_BORDER
        _write_kpi_row(ws4, r, row, 2)

        spend_pct = row.get("spend", 0) / total_spend if total_spend else 0
        pct_cell = ws4.cell(row=r, column=12, value=spend_pct)
        pct_cell.number_format = FMT_PCT
        pct_cell.font = BODY_FONT
        pct_cell.border = THIN_BORDER

    _auto_width(ws4)

    # ============================================================
    # Sheet 5: 广告活动 × 展示位置交叉分析（数据透视）
    # ============================================================
    ws5 = wb.create_sheet("交叉分析(花费)")
    ws5.sheet_properties.tabColor = "EF4444"

    # Build pivot: campaign × placement → spend
    from backend.models import PlacementRecord, Campaign
    from sqlalchemy import func

    q = (
        db.query(
            Campaign.name,
            PlacementRecord.placement_type,
            func.sum(PlacementRecord.spend),
            func.sum(PlacementRecord.orders),
        )
        .join(Campaign)
        .group_by(Campaign.name, PlacementRecord.placement_type)
    )

    if date_from:
        q = q.filter(PlacementRecord.date >= date_from)
    if date_to:
        q = q.filter(PlacementRecord.date <= date_to)

    pivot_data = q.all()

    # Collect unique placements and campaigns
    placements = sorted(set(r[1] for r in pivot_data))
    campaigns = sorted(set(r[0] for r in pivot_data))

    # Header row
    ws5.cell(row=1, column=1, value="广告活动 \\ 展示位置").font = HEADER_FONT
    ws5.cell(row=1, column=1).fill = HEADER_FILL
    ws5.cell(row=1, column=1).border = THIN_BORDER
    for j, p in enumerate(placements):
        cell = ws5.cell(row=1, column=2 + j, value=p)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
    # Total column
    total_col = 2 + len(placements)
    cell = ws5.cell(row=1, column=total_col, value="合计")
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.border = THIN_BORDER

    # Build lookup
    spend_map = {}
    for name, ptype, spend, orders in pivot_data:
        spend_map[(name, ptype)] = (spend or 0, orders or 0)

    for i, camp in enumerate(campaigns):
        r = i + 2
        ws5.cell(row=r, column=1, value=camp).font = BODY_FONT
        ws5.cell(row=r, column=1).border = THIN_BORDER
        row_total = 0
        for j, p in enumerate(placements):
            spend, orders = spend_map.get((camp, p), (0, 0))
            cell = ws5.cell(row=r, column=2 + j, value=round(spend, 2))
            cell.number_format = FMT_CURRENCY
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            row_total += spend
        # Row total formula
        first_col = get_column_letter(2)
        last_col = get_column_letter(1 + len(placements))
        cell = ws5.cell(row=r, column=total_col)
        cell.value = f"=SUM({first_col}{r}:{last_col}{r})"
        cell.number_format = FMT_CURRENCY
        cell.font = Font(name="Arial", bold=True, size=10)
        cell.border = THIN_BORDER

    # Column totals
    total_row = len(campaigns) + 2
    ws5.cell(row=total_row, column=1, value="合计").font = Font(name="Arial", bold=True, size=10)
    ws5.cell(row=total_row, column=1).border = THIN_BORDER
    for j in range(len(placements) + 1):
        col = 2 + j
        col_letter = get_column_letter(col)
        cell = ws5.cell(row=total_row, column=col)
        cell.value = f"=SUM({col_letter}2:{col_letter}{total_row - 1})"
        cell.number_format = FMT_CURRENCY
        cell.font = Font(name="Arial", bold=True, size=10)
        cell.border = THIN_BORDER

    _auto_width(ws5)

    # ============================================================
    # Sheet 6: 预警与建议
    # ============================================================
    ws6 = wb.create_sheet("预警与建议")
    ws6.sheet_properties.tabColor = "DC2626"

    alerts = overview.get("alerts", [])
    ws6.append(["严重级别", "广告活动", "指标值", "建议"])
    _style_header(ws6, 1, 4)

    for alert in alerts:
        r = ws6.max_row + 1
        ws6.cell(row=r, column=1, value=alert.get("severity", "")).font = BODY_FONT
        ws6.cell(row=r, column=1).border = THIN_BORDER
        ws6.cell(row=r, column=2, value=alert.get("campaign_name", "")).font = BODY_FONT
        ws6.cell(row=r, column=2).border = THIN_BORDER
        ws6.cell(row=r, column=3, value=str(alert.get("value", ""))).font = BODY_FONT
        ws6.cell(row=r, column=3).border = THIN_BORDER
        ws6.cell(row=r, column=4, value=alert.get("message", "")).font = BODY_FONT
        ws6.cell(row=r, column=4).border = THIN_BORDER

        if alert.get("severity") in ("danger", "warning"):
            for col in range(1, 5):
                ws6.cell(row=r, column=col).fill = BAD_FILL

    if not alerts:
        ws6.append(["", "所有广告活动运行正常，暂无预警", "", ""])

    _auto_width(ws6)

    # Write to bytes
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
